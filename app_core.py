import os
import io
import time
import platform
import subprocess
import re
import cv2
import secrets
import base64
import uuid
import numpy as np
from pdf_extractor import process_pdf_upload
from datetime import datetime, timedelta, timezone, date

# ── IST helper (UTC+5:30) — syncs from NTP, never depends on local clock ────
# Moved to time_utils.py so ocr_mrz.py can use the same "today" for MRZ
# date-of-birth century inference without importing app.py.
from time_utils import ist_now, sync_ntp_async
# ─────────────────────────────────────────────────────────────────────────────
import zipfile
from cloudstore_backup import upload_zip_to_cloudstore
from flask import (
    Flask, request, render_template,
    redirect, url_for, session,
    send_file, send_from_directory, flash, jsonify, make_response, render_template_string, g
)
from werkzeug.utils import secure_filename
from PIL import Image, ImageOps, ImageDraw, ImageFont
from dotenv import load_dotenv
from functools import wraps

load_dotenv()

# ── Module-level constants (computed once, reused everywhere) ────────────────
try:
    _RESAMPLE = Image.Resampling.LANCZOS
except AttributeError:
    _RESAMPLE = Image.LANCZOS  # Pillow < 9.1

import sys
def _resource_path(rel):
    base = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base, rel)

def _find_cascade():
    # 1. Bundled next to exe / script (--add-data "haarcascade...;.")
    p1 = _resource_path("haarcascade_frontalface_default.xml")
    if os.path.exists(p1):
        return p1
    # 2. OpenCV's own built-in data folder (fallback)
    try:
        p2 = os.path.join(os.path.dirname(cv2.__file__), "data", "haarcascade_frontalface_default.xml")
        if os.path.exists(p2):
            return p2
    except Exception:
        pass
    return p1  # return anyway, CascadeClassifier will handle missing gracefully

_CASCADE_PATH = _find_cascade()
_face_cascade = cv2.CascadeClassifier(_CASCADE_PATH)  # loaded once, reused per request
# ─────────────────────────────────────────────────────────────────────────────

from config import (
    UPLOAD_FOLDER,
    FACE_FOLDER,
    DB_CONFIG,
    BASE_DIR
)

# NOTE: ProvA/ProvB Vision credentials (PROVA_ENDPOINT, PROVA_KEY,
# CLOUDSTORE_APPLICATION_CREDENTIALS) are no longer imported or used here. OCR
# scanning now happens entirely on the VPS (see ocr_scan_service.py) — this
# local process only crops the MRZ strip and sends it over HTTPS via
# ocr_client.py. config.py still decrypts the CLOUDSTORE service-account file at
# startup regardless (untouched, per its own module), but the result is no
# longer consumed anywhere in this app.

from db import (
    init_db,
    get_connection,
    insert_passport,
    insert_general_data,
    copy_passport_files_to_group,
    safe_int,
    get_active_passport_in_group,
    is_passport_number_exists_in_group,
    is_passport_number_exists_in_group_same_group_only,
    get_group_visa_type,
    find_duplicate_groups_for_passports,
    get_all_passports_with_general_data,
    get_total_passport_count,
    get_passport_by_id,
    get_passport_by_number,
    update_passport_data,
    update_general_data,
    delete_passport_record,
    insert_invalid_passport,
    get_invalid_passport_by_id,
    get_all_invalid_passports,
    delete_invalid_passport,
    get_total_invalid_count,
    get_user_settings,
    update_user_settings,
    get_filename_group_map,
    get_all_recycled_invalid_passports,
    get_total_recycled_invalid_count,
    restore_invalid_passport,
    hard_delete_invalid_passport,
    auto_empty_recycle_bin,
    auto_remove_old_invalid_records,
    get_recycled_passports,
    get_total_recycled_passports_count,
    restore_passport,
    hard_delete_passport,
    empty_recycled_passports,
    DuplicateRestoreConflict,
    PLAN_DETAILS,
    get_all_group_batches,
    touch_group_activity,
    upsert_group_batch,
    create_visit_visa_queue,
    pop_next_visit_visa_queue_item,
    mark_visit_visa_queue_item_finished,
    get_visit_visa_queue_user,
    get_visit_visa_queue_user_any_status,
    clear_visit_visa_queue_for_user,
    delete_group_and_records,
    rename_group,
    set_progress,
    update_progress_field,
    get_progress,
    clear_progress,
    get_passport_path,
    resolve_passport_paths,
    get_visa_processed_group_dir,
    list_visa_processed_pdfs,
    get_visa_pdf_path,
    move_group_folder,
    move_passport_files_to_group,
    move_passport_to_group,
    move_passports_to_group,
    create_nusuk_queue,
    pop_next_nusuk_queue_item,
    mark_nusuk_queue_item_finished,
    clear_nusuk_queue_for_user,
    get_nusuk_queue_user,
    get_active_nusuk_queue_user,
    register_visit_visa_exe,
    get_visit_visa_exe_registration,
    touch_visit_visa_exe_registration,
    delete_visit_visa_exe_registration,
    reconcile_mofa_pdf_downloads
)

from host_api import (
    host_login,
    host_check_upload,
    host_log_upload,
    host_account_summary,
    host_change_credentials,
)


def report_usage_to_host(user_id, count=1, duplicates=0, invalids=0):
    """
    Mirror an upload/correction usage count to the host's quota/log.php
    (via host_log_upload) so server-side used-limit and daily-usage stats
    stay in sync with the local DB's log_passport_upload() calls.

    Every call site that logs locally should call this alongside it.
    Silently skipped (not raised) if there's no host_token in the current
    session -- e.g. the host session expired -- so a host-auth issue never
    blocks the local upload/correction flow that already succeeded.
    """
    host_token = session.get('host_token')
    if not host_token:
        return
    host_log_upload(user_id, host_token, count=count, duplicates=duplicates, invalids=invalids)

from mrz_utils import crop_passport_face

# parse_mrz now makes a network call to the VPS (mrz_core.py, via
# /ocr/parse_mrz) instead of running the checksum/regex logic locally —
# moved server-side so passport MRZ validation rules aren't shipped
# inside the distributable exe. Same signature/return contract as the
# old local mrz_utils.parse_mrz, so every call site below is unchanged.
#
# check_issue_date_rule() moved alongside it for the same reason (and
# for consistency with convert_mrz_date/estimate_issue_date below,
# which already made this same move).
from ocr_client import parse_mrz, check_issue_date_rule

# All OCR scanning + MRZ assembly/reparsing logic now runs on the VPS
# (see ocr_scan_service.py). This local process only crops the MRZ strip
# out of the passport image itself (needs the full local file) and sends
# that small strip over HTTPS to the VPS for the actual ProvA/ProvB Vision
# calls + assembly — see ocr_client.py for the local-side half of that split.
from ocr_client import (
    BATCH_SIZE,
    primary_client,
    _reset_secondary_client,
    _is_token_expired_error,
    extract_mrz_from_image,
    extract_mrz_from_image_crop_rescan,
    extract_mrz_from_image_reparse_rescan,
    extract_mrz_from_region_reparse_rescan,
    _do_stacked_primary_scan_from_paths,
    _do_stacked_secondary_scan,
    convert_mrz_date,
    estimate_issue_date,
    get_last_extracted_issue_date,
)

# Static nationality/marital-status lookup tables + the Nusuk group sort helper.
from constants import (
    NATIONALITY_CODE_MAP,
    NATIONALITY_ID_TO_COUNTRY_CODE,
    NATIONALITY_OPTIONS,
    MARITAL_STATUS_OPTIONS,
    sort_nusuk_group,
)

import mysql.connector
import json
import csv
import openpyxl
import threading
import shutil

# =====================================================
# APP INIT
# =====================================================

app = Flask(__name__, template_folder=_resource_path("templates"), static_folder=_resource_path("static"))


class ForceHTTPS(object):
    def __init__(self, app):
        self.app = app

    def __call__(self, environ, start_response):
        environ['wsgi.url_scheme'] = 'https'
        return self.app(environ, start_response)


app.wsgi_app = ForceHTTPS(app.wsgi_app)

import logging
import traceback as _traceback
_logger = logging.getLogger('passposys')

@app.errorhandler(Exception)
def _log_unhandled_exception(e):
    """
    Catch-all for any unhandled exception in any route. Logs the full
    traceback to passposys.log so a bare 500 page never happens without
    a corresponding entry in the log file. Re-raises so Flask's normal
    error page/behavior is unchanged for the person using the app.
    """
    _logger.error(
        "Unhandled exception on %s %s:\n%s",
        request.method, request.path,
        ''.join(_traceback.format_exception(type(e), e, e.__traceback__))
    )
    raise e

@app.before_request
def _refresh_ntp_on_request():
    """Re-sync NTP offset in background on every page load.
    Runs in a daemon thread so it never slows down the response."""
    try:
        sync_ntp_async()
    except Exception:
        _logger.exception("sync_ntp_async() failed to start - continuing without NTP resync")

# ProvA/ProvB Vision clients now live in ocr_mrz.py (imported near the top
# of this file, alongside the other OCR/MRZ names) so there is a single
# client instance shared by every OCR call site.

app.secret_key = os.environ.get('FLASK_SECRET_KEY')
# LOCAL_API_TOKEN must be set via environment variable (env.enc) — never hardcoded.
LOCAL_API_TOKEN = os.environ.get('LOCAL_API_TOKEN', '')
if not LOCAL_API_TOKEN:
    raise RuntimeError(
        "LOCAL_API_TOKEN is not set. Add it to env.enc or set the LOCAL_API_TOKEN environment variable."
    )
# Write token to a shared file so the standalone exe can read it without
# relying on env-var inheritance (the exe is often launched independently,
# not as a child of this Flask process, so it never gets Flask's env vars).
try:
    with open(os.path.join(BASE_DIR, "local_api_token.dat"), "w") as _f:
        _f.write(LOCAL_API_TOKEN)
except Exception:
    _logger.exception("Could not write local_api_token.dat — exe DB updates may fail")
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
app.config["FACE_FOLDER"] = FACE_FOLDER
# No upload size limit (local system)

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(FACE_FOLDER, exist_ok=True)

# init_db() runs on every startup (gunicorn or direct). It is fully idempotent:
# uses CREATE TABLE IF NOT EXISTS and SHOW COLUMNS + ALTER TABLE guards, so it
# only adds missing columns/tables and never damages existing data.
with app.app_context():
    init_db()

# Scheduled auto-backup (dump DB + images -> zip -> keep last 2 locally ->
# upload to CloudStore) now lives in backup.py; only one gunicorn worker
# actually runs the schedule (guarded by a lock file inside start_scheduler).
from backup import _find_mysqldump, LOCAL_BACKUP_DIR, BACKUP_TEMP_DIR, start_scheduler
start_scheduler(app)





# =====================================================
# CANCEL / ROLLBACK STATE
# Per-user cancel flags. Keyed by user_id (int).
# Set to True by /api/cancel when the browser sends a
# sendBeacon on pagehide. Checked between processing
# phases and between files in Phase 5.
# =====================================================
import threading as _cancel_thr
_cancel_flags      = {}          # {user_id: True}
_cancel_flags_lock = _cancel_thr.Lock()


# Session-aware quota/usage-warning helpers now live in quota.py.
from quota import (
    get_plan_usage_summary,
    check_upload_allowed,
    log_passport_upload,
    _set_usage_warning,
)


def _rollback_session(user_id, passport_ids, invalid_ids, file_paths):
    """Hard-delete all DB records and disk files saved during a cancelled upload.

    Deletes rows directly (no soft-delete / recycle bin) because the user
    never saw these records — they were inserted during the chunk that was
    interrupted by the refresh.
    """
    try:
        if passport_ids:
            conn = get_connection()
            cursor = conn.cursor()
            cursor.execute("USE passport_db")
            ph = ','.join(['%s'] * len(passport_ids))
            # CASCADE deletes the matching general_data rows automatically.
            cursor.execute(
                f"DELETE FROM passports WHERE id IN ({ph}) AND user_id = %s",
                passport_ids + [user_id]
            )
            conn.commit()
            cursor.close()
            conn.close()
            print(f"[Cancel] Rolled back {len(passport_ids)} passport record(s) for user {user_id}.")
    except Exception as _e:
        print(f"[Cancel] Rollback passport records failed: {_e}")

    try:
        if invalid_ids:
            conn = get_connection()
            cursor = conn.cursor()
            cursor.execute("USE passport_db")
            ph = ','.join(['%s'] * len(invalid_ids))
            cursor.execute(
                f"DELETE FROM invalid_passports WHERE id IN ({ph}) AND user_id = %s",
                invalid_ids + [user_id]
            )
            conn.commit()
            cursor.close()
            conn.close()
            print(f"[Cancel] Rolled back {len(invalid_ids)} invalid record(s) for user {user_id}.")
    except Exception as _e:
        print(f"[Cancel] Rollback invalid records failed: {_e}")

    for _path in file_paths:
        try:
            if _path and os.path.exists(_path):
                os.remove(_path)
        except Exception as _e:
            print(f"[Cancel] Could not delete file {_path}: {_e}")


def read_blob(path):
    try:
        with open(path, "rb") as f:
            return f.read()
    except Exception as e:
        print(f"Error reading file {path}: {e}")
        return None


def compress_image_bytes_for_disk(raw_blob, max_bytes=1_000_000):
    """Shrink an image below max_bytes (default 1 MB) before it's written to
    the filesystem. Images already under max_bytes are returned unchanged
    (no quality loss). Over the limit: halve dimensions, then step down JPEG
    quality until the result fits under max_bytes. Applies to both valid and
    invalid passport uploads, since both are written to disk via the same
    code path. Non-image bytes (or anything Pillow can't open) are returned
    unchanged so we never corrupt/drop a file we can't process."""
    if not raw_blob or len(raw_blob) <= max_bytes:
        return raw_blob or b""
    try:
        with Image.open(io.BytesIO(raw_blob)) as img:
            img = ImageOps.exif_transpose(img)
            if img.mode != 'RGB':
                img = img.convert('RGB')

            w, h = img.size
            img = img.resize((w // 2, h // 2), _RESAMPLE)

            compressed = raw_blob
            for quality in [85, 70, 55, 40, 25]:
                buf = io.BytesIO()
                img.save(buf, format='JPEG', quality=quality)
                compressed = buf.getvalue()
                if len(compressed) <= max_bytes:
                    return compressed

            # Last resort: whatever quality=25 produced, even if still >1MB
            return compressed
    except Exception as _e:
        print(f"  ⚠️  compress_image_bytes_for_disk failed, keeping original: {_e}")
        return raw_blob


def _compress_blob_for_db(raw_blob, max_bytes=1_000_000, target_bytes=500_000):
    """Store image in DB. If the raw blob exceeds max_bytes (1 MB), halve the
    dimensions first and then compress to ~500 KB. Images already under 1 MB
    are stored as-is (no quality loss)."""
    if not raw_blob:
        return b""
    try:
        with Image.open(io.BytesIO(raw_blob)) as img:
            img = ImageOps.exif_transpose(img)
            if img.mode != 'RGB':
                img = img.convert('RGB')

            # Under 1 MB — store original bytes unchanged
            if len(raw_blob) <= max_bytes:
                return raw_blob

            # Over 1 MB — halve dimensions first
            w, h = img.size
            img = img.resize((w // 2, h // 2), _RESAMPLE)

            # Then compress down to ~500 KB
            for quality in [85, 70, 55, 40, 25]:
                buf = io.BytesIO()
                img.save(buf, format='JPEG', quality=quality)
                compressed = buf.getvalue()
                if len(compressed) <= target_bytes:
                    return compressed

            # Last resort: return whatever quality=25 produced
            return compressed

    except Exception as _e:
        print(f"  ⚠️  _compress_blob_for_db failed: {_e}")
        return b""

def save_invalid_to_db(user_id, filename, image_source, mrz_text, error_message,
                        upload_group_name=None, upload_visa_type=None, extracted_issue_date=None,
                        is_emergency=False):
    if error_message:
        err_lower = error_message.lower()
        if any(kw in err_lower for kw in ['provA', 'visionsdk', '503', 'vision', 'api', 'service unavailable']):
            group_addon = ""
            if "(Group:" in error_message:
                try:
                    group_addon = " (Group: " + error_message.split("(Group: ")[1]
                except Exception:
                    pass
            error_message = f"Can't extract due to server error{group_addon}"

    blob = b""
    if isinstance(image_source, str) and image_source and os.path.exists(image_source):
        blob = read_blob(image_source) or b""
    elif isinstance(image_source, (bytes, bytearray)) and image_source:
        blob = bytes(image_source)

    safe_error_msg = (error_message[:250] + '...') if error_message and len(error_message) > 250 else error_message

    for attempt, use_blob in enumerate([blob, b""]):
        try:
            inv_id = insert_invalid_passport(
                user_id=user_id,
                filename=filename,
                original_blob=use_blob,
                mrz_text=mrz_text or "",
                error_message=safe_error_msg,
                upload_group_name=upload_group_name,
                upload_visa_type=upload_visa_type,
                extracted_issue_date=extracted_issue_date,
                is_emergency=is_emergency
            )
            
            if not inv_id:
                raise ValueError("DB insertion silently returned None.")

            # NOTE: usage is no longer reported here (per-item host call removed
            # for performance). The caller aggregates invalid counts and reports
            # them once at the end of the batch via log_passport_upload().

            return inv_id

        except Exception as _db_err:
            print(f"  ❌ Invalid insert attempt {attempt + 1} failed: {_db_err}")

    return None


# =====================================================
# OCR scanning + MRZ assembly/reparsing logic has moved to ocr_mrz.py.
# See the `from ocr_client import (...)` block near the top of this file for
# everything that's still used here (extract_mrz_from_image*, convert_mrz_date,
# estimate_issue_date, the stacked-scan batch helpers, and the shared Vision
# clients / token-expiry helpers).
# =====================================================


# =====================================================
# IMAGE SERVING FROM DATABASE (CACHE-BUSTED)
# =====================================================

# How often (seconds) login_required() re-verifies the install_token
# against the host, per session, rather than on literally every single
# request (avoids a host round-trip on every AJAX poll/static asset
# fetch while a page is open). The session itself remembers the last
# check time, so this is still "checked on every request" in effect --
# it just doesn't re-hit the host API more often than this interval.
_INSTALL_TOKEN_CHECK_INTERVAL_SECONDS = 60


def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('user_id'):
            return redirect(url_for('login', next=request.path))

        last_checked = session.get('_install_token_checked_at', 0)
        now_ts = time.time()
        if now_ts - last_checked >= _INSTALL_TOKEN_CHECK_INTERVAL_SECONDS:
            try:
                from secrets_client import check_install_token_valid
                result = check_install_token_valid(session['user_id'])
            except Exception:
                _logger.exception("login_required: install_token check raised unexpectedly; continuing session.")
                result = None

            if result is False:
                # Token is missing locally, revoked, or belongs to a
                # different user than this session -- the server no
                # longer recognizes this device as authorized for this
                # account. Force logout rather than let the session
                # silently continue.
                _logger.warning(
                    "login_required: install_token mismatch for user_id=%s -- forcing logout.",
                    session.get('user_id')
                )
                session.clear()
                flash(
                    'Your device could not be verified for this account '
                    '(install token missing or invalid). You have been logged out '
                    'for security. Please contact your administrator if this persists.',
                    'danger'
                )
                return redirect(url_for('login'))

            # result is True (verified) or None (check failed/unreachable
            # -- don't penalize the user for a transient network issue).
            # Either way, only advance the timer on a completed attempt
            # so a transient failure gets retried on the very next
            # request rather than waiting out the full interval.
            if result is not None:
                session['_install_token_checked_at'] = now_ts

        return f(*args, **kwargs)
    return decorated_function


# Fields from the "General Data Defaults" card that are bulk-applied to
# every existing record in a group when the Update button is used.
def _resolve_default_arrival_departure(db_defaults, now, one_year_later):
    """
    Returns (expected_arrival_date, expected_departure_date) for new records.
    Uses the admin-configured default arrival date (set on the index page)
    when present, otherwise falls back to today. Departure is always
    arrival + 365 days — it isn't stored independently, since the index
    page auto-recalculates it whenever the arrival date changes.
    (Nusuk records ignore both of these — insert_general_data() leaves
    them NULL regardless of what's passed in here.)
    """
    raw_arrival = db_defaults.get('expected_arrival') if db_defaults else None
    arrival_date = None
    if raw_arrival:
        if isinstance(raw_arrival, date):
            arrival_date = raw_arrival
        else:
            try:
                arrival_date = datetime.strptime(str(raw_arrival), "%Y-%m-%d").date()
            except (ValueError, TypeError):
                arrival_date = None
    if arrival_date:
        return arrival_date, arrival_date + timedelta(days=365)
    return now.date(), one_year_later.date()


GROUP_BULK_UPDATE_FIELDS = [
    'marital_status', 'city_of_birth', 'profession', 'city',
    'zip_postal_code', 'address', 'passport_issue_place', 'hotel_name',
    'email', 'contact_number', 'passport_type', 'visa_type',
    'expected_arrival',
]


def get_distinct_group_names():
    try:
        if not session.get('user_id'): return []
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("USE passport_db")
        cursor.execute("""
            SELECT gd.group_name
            FROM general_data gd
            JOIN passports p ON gd.passport_id = p.id
            WHERE gd.group_name IS NOT NULL AND gd.group_name != '' AND p.user_id = %s
            GROUP BY gd.group_name
            ORDER BY MAX(p.created_at) DESC
        """, (session['user_id'],))
        rows = cursor.fetchall()
        cursor.close()
        conn.close()
        return [row[0] for row in rows]
    except Exception as e:
        print(f"Error fetching group names: {e}")
        return []


def get_distinct_group_visa_types():
    """
    Returns {group_name: 'nusuk' | 'visit_visa'} for every existing group of
    the current user. Used by the Upload page so the "Group Already Exists"
    choice modal can block "Add applicant to the existing group" client-side
    when the currently-selected visa type doesn't match what that group is
    already locked to — instead of silently letting a mixed-type record
    through, which the server also rejects (see _assert_group_visa_type_matches).
    """
    try:
        if not session.get('user_id'):
            return {}
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("USE passport_db")
        cursor.execute("""
            SELECT gd.group_name, COALESCE(MAX(gd.visa_type), 'nusuk')
            FROM general_data gd
            JOIN passports p ON gd.passport_id = p.id
            WHERE gd.group_name IS NOT NULL AND gd.group_name != '' AND p.user_id = %s
              AND (p.is_recycled = FALSE OR p.is_recycled IS NULL)
            GROUP BY gd.group_name
        """, (session['user_id'],))
        rows = cursor.fetchall()
        cursor.close()
        conn.close()
        return {row[0]: (row[1] or 'nusuk') for row in rows}
    except Exception as e:
        print(f"Error fetching group visa types: {e}")
        return {}


def get_distinct_group_emergency_flags():
    """
    Returns {group_name: True/False} for every existing group of the
    current user — True if ANY record in that group was saved while the
    "Emergency upload" checkbox was checked. Used to color the group red
    in the upload-page dropdown and to show an "Emergency" badge on the
    Groups page.
    """
    try:
        if not session.get('user_id'):
            return {}
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("USE passport_db")
        cursor.execute("""
            SELECT gd.group_name, MAX(gd.is_emergency)
            FROM general_data gd
            JOIN passports p ON gd.passport_id = p.id
            WHERE gd.group_name IS NOT NULL AND gd.group_name != '' AND p.user_id = %s
              AND (p.is_recycled = FALSE OR p.is_recycled IS NULL)
            GROUP BY gd.group_name
        """, (session['user_id'],))
        rows = cursor.fetchall()
        cursor.close()
        conn.close()
        return {row[0]: bool(row[1]) for row in rows}
    except Exception as e:
        print(f"Error fetching group emergency flags: {e}")
        return {}


# =====================================================
# MAIN UPLOAD ROUTE  ──  4-PHASE OCR PIPELINE
# =====================================================
#
#  PHASE 1  ─ Rotation validation (OpenCV, zero API cost)
#             ↳ Rotated images  → saved as invalid, skipped
#             ↳ Valid images    → added to pending_items
#
#  PHASE 2  ─ ProvB Stacked Batch scan
#             ↳ All pending_items split into batches of 13, stacked into one
#               tall image per batch, sent to ProvB Vision OCR.
#             ↳ Success                → MRZ saved for that item
#             ↳ Server fail (ProvB)   → try ProvA stacked for that same
#                                          batch; if that ALSO fails → invalid
#             ↳ After this phase:
#                 • > 3 invalid items   → continue to Phase 3
#                 • <= 3 invalid items  → save remaining as invalid, stop
#
#  PHASE 3  ─ ProvA Stacked Batch scan
#             ↳ Phase-2 failures split into batches of 13, stacked, sent to
#               ProvA Vision Read API.
#             ↳ Success                → MRZ saved for that item
#             ↳ Server fail (ProvA)    → try ProvB stacked for that same
#                                          batch; if that ALSO fails → invalid
#             ↳ After this phase:
#                 • > 3 invalid items   → continue to Phase 4
#                 • <= 3 invalid items  → save remaining as invalid, stop
#
#  PHASE 4  ─ Individual ProvA scan
#             ↳ Each Phase-3 failure scanned individually by ProvA.
#             ↳ Success                → MRZ saved for that item
#             ↳ Server fail (ProvA)    → try ProvB individual scan; if that
#                                          ALSO fails → invalid
#             ↳ Still fails            → save as invalid, stop
#
#  PHASE 5  ─ Final DB processing
#             ↳ Valid MRZ → insert to passports + general_data
#             ↳ Still-failed → insert to invalid_passports
# =====================================================

def _extract_passports(filtered=False, passport_number=None, group_name=None):
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("USE passport_db")

    where_clauses = ["p.user_id = %s", "(p.is_recycled = FALSE OR p.is_recycled IS NULL)"]
    params = [session['user_id']]

    if filtered:
        if passport_number:
            where_clauses.append("p.passport_number LIKE %s")
            params.append(f"%{passport_number}%")
        if group_name:
            where_clauses.append("g.group_name = %s")
            params.append(group_name)

    where_sql = " WHERE " + " AND ".join(where_clauses)
    query = f"""
        SELECT p.id, p.passport_number, p.filename, g.group_name
        FROM passports p
        LEFT JOIN general_data g ON p.id = g.passport_id
        {where_sql}
        ORDER BY p.created_at DESC
    """
    cursor.execute(query, tuple(params))
    passports_meta = cursor.fetchall()

    memory_file = io.BytesIO()
    seen_names = set()

    with zipfile.ZipFile(memory_file, 'w') as zf:
        for meta in passports_meta:
            img_filename = meta.get('filename')
            if not img_filename:
                continue
            img_path, _ = resolve_passport_paths(img_filename, meta.get('group_name'))
            if not img_path or not os.path.exists(img_path):
                continue
            pn = str(meta.get('passport_number', '')).strip() or f"unknown_id_{meta['id']}"
            out_name = f"{pn}.jpg"
            counter = 1
            while out_name in seen_names:
                out_name = f"{pn}_{counter}.jpg"
                counter += 1
            seen_names.add(out_name)
            zf.write(img_path, out_name)

    cursor.close()
    conn.close()
    memory_file.seek(0)

    return send_file(
        memory_file,
        mimetype='application/zip',
        as_attachment=True,
        download_name=f'passports_export_{ist_now().strftime("%Y%m%d_%H%M%S")}.zip'
    )


class GroupVisaTypeMismatch(Exception):
    """Raised when trying to insert a record whose visa type doesn't match
    the visa type already locked in by the target group."""
    pass


def _assert_group_visa_type_matches(user_id, group_name, visa_type):
    """Shared guard used by every insert path (normal reparse, insert-anyway,
    and force-insert) so an existing group can never end up holding a mix
    of Nusuk and Visit Visa records, regardless of which route the record
    came in through."""
    _visa_label = {'nusuk': 'Nusuk', 'visit_visa': 'Visit Visa'}
    _existing = get_group_visa_type(user_id, group_name)
    if _existing and _existing != visa_type:
        raise GroupVisaTypeMismatch(
            f'Group "{group_name}" is already a '
            f'{_visa_label.get(_existing, _existing)} group. '
            f'You are trying to add a {_visa_label.get(visa_type, visa_type)} '
            f'record — only matching visa types can be added to the same group.'
        )


def process_valid_invalid_passport(invalid_id, parsed, mrz_text, form_data):
    user_id = session['user_id']
    try:
        _pvip_settings = get_user_settings(user_id) or {}
        _assert_group_visa_type_matches(
            user_id,
            _pvip_settings.get('group_name', 'GROUP 1'),
            _pvip_settings.get('visa_type', 'nusuk')
        )
        invalid_passport = get_invalid_passport_by_id(invalid_id, user_id)
        passport_blob = invalid_passport['original_image']
        inv_filename = invalid_passport['filename']

        # Write original image to permanent UPLOAD_FOLDER (filesystem storage)
        perm_orig_path = os.path.join(app.config["UPLOAD_FOLDER"], inv_filename)
        perm_face_path = os.path.join(app.config["FACE_FOLDER"], f"face_{inv_filename}")
        temp_work_path = os.path.join(app.config["UPLOAD_FOLDER"], f"_work_{invalid_id}.jpg")

        try:
            # Save the ORIGINAL (EXIF-corrected, unrotated) image straight to its
            # permanent location — this is what gets displayed on the results page
            # and must never end up rotated, regardless of which orientation the
            # face-detection loop below ends up using.
            try:
                with Image.open(io.BytesIO(passport_blob)) as _orig_img:
                    _orig_img = ImageOps.exif_transpose(_orig_img)
                    if _orig_img.mode != 'RGB':
                        _orig_img = _orig_img.convert('RGB')
                    _orig_img.save(perm_orig_path, format='JPEG', quality=95)
            except Exception:
                # Fallback: write the raw blob as-is if it can't be opened/re-encoded
                with open(perm_orig_path, "wb") as f:
                    f.write(passport_blob)

            # Separate working copy used only to search for a face crop at
            # different rotations — never written back to perm_orig_path.
            with open(temp_work_path, "wb") as f:
                f.write(passport_blob)

            face_found = False
            try:
                for _fa in range(4):
                    if _fa > 0:
                        with Image.open(temp_work_path) as _fi:
                            _fi = ImageOps.exif_transpose(_fi)
                            if _fi.mode != 'RGB':
                                _fi = _fi.convert('RGB')
                            _fr = _fi.rotate(90, expand=True)
                            _fb = io.BytesIO()
                            _fr.save(_fb, format='JPEG', quality=95)
                            with open(temp_work_path, 'wb') as _fw:
                                _fw.write(_fb.getvalue())
                    if crop_passport_face(temp_work_path, perm_face_path):
                        face_found = True
                        break
            except Exception as face_err:
                print(f"Face crop error during reparse (image will still be saved): {face_err}")
            finally:
                if os.path.exists(temp_work_path):
                    os.remove(temp_work_path)
        except Exception as e:
            print(f"File error during reparse: {e}")
            if os.path.exists(temp_work_path):
                try: os.remove(temp_work_path)
                except: pass

        # Insert passport record (no BLOBs — images on filesystem)
        passport_id = insert_passport(
            user_id, parsed, None, None,
            mrz_text, inv_filename
        )

        now = ist_now()
        one_year_later = now + timedelta(days=365)
        mrz_nationality_id = NATIONALITY_CODE_MAP.get(parsed.get("nationality"), 197)

        manual_issue = form_data.get("manual_issue_date")
        if manual_issue:
            try:
                final_issue_date = datetime.strptime(manual_issue, "%Y-%m-%d").date()
            except ValueError:
                # Manual entry was malformed -- this flow has no raw OCR
                # text to extract from (manual reparse page, typed-in MRZ),
                # so we never estimate here. passport_issue_date is a
                # nullable column, so store NULL rather than guessing a
                # placeholder date; the user can fill it in afterward.
                final_issue_date = None
        else:
            # No manual override supplied. This reparse path has no raw OCR
            # text of its own to extract from, but the original main-scan
            # pass (when this record was first filed as invalid) may have
            # already captured an OCR-extracted issue date -- reuse that
            # instead of giving up. Still never estimate/guess: if that
            # column is also NULL, final_issue_date stays NULL.
            final_issue_date = invalid_passport.get('extracted_issue_date')

        db_defaults = get_user_settings(user_id)
        _arr_date, _dep_date = _resolve_default_arrival_departure(db_defaults, now, one_year_later)
        _group_for_save = db_defaults.get('group_name', 'GROUP 1')

        # Move the permanently-written images into the group's own folder.
        _final_orig_path = get_passport_path(inv_filename, _group_for_save, kind="original")
        try:
            if os.path.exists(perm_orig_path) and perm_orig_path != _final_orig_path:
                shutil.move(perm_orig_path, _final_orig_path)
        except Exception as _move_e:
            print(f"  ⚠️ Could not move reparsed original into group folder: {_move_e}")
        _final_face_path = get_passport_path(inv_filename, _group_for_save, kind="face")
        try:
            if os.path.exists(perm_face_path) and perm_face_path != _final_face_path:
                shutil.move(perm_face_path, _final_face_path)
        except Exception as _move_e:
            print(f"  ⚠️ Could not move reparsed face into group folder: {_move_e}")

        insert_general_data(
            passport_id=passport_id,
            nationality_id=mrz_nationality_id,
            marital_status=safe_int(db_defaults.get('marital_status'), 5),
            group_name=_group_for_save,
            city_of_birth=db_defaults.get('city_of_birth', 'MAIN STREET'),
            profession=db_defaults.get('profession', 'TOURISM'),
            city=db_defaults.get('city', 'MAIN STREET'),
            zip_postal_code=db_defaults.get('zip_postal_code', '676542'),
            address=db_defaults.get('address', 'ADDRESS'),
            passport_type=safe_int(db_defaults.get('passport_type'), 1),
            passport_issue_place=db_defaults.get('passport_issue_place', 'PLACE'),
            passport_issue_date=final_issue_date,
            expected_arrival=_arr_date,
            expected_departure=_dep_date,
            hotel_name=db_defaults.get('hotel_name', 'Hayat Mall Gate 6, Riyadh'),
            contact_number=db_defaults.get('contact_number', ''),
            email=db_defaults.get('email', ''),
            visa_type=db_defaults.get('visa_type', 'nusuk'),
            is_emergency=bool(invalid_passport.get('is_emergency'))
        )

        hard_delete_invalid_passport(invalid_id, user_id)
        # log_passport_upload() already reports to the host internally via
        # host_log_upload() — do not also call report_usage_to_host() here,
        # it duplicates the count against the server-side quota.
        log_passport_upload(user_id, count=1, duplicates=0, invalids=-1)

        remaining = get_total_invalid_count(user_id)
        parent_redirect = url_for('results') if remaining == 0 else url_for('view_invalid_passports')

        return f'''
        <!DOCTYPE html><html><head><title>Success</title>
        <script>
            if (window.opener && !window.opener.closed && !window.opener._reparseActive) {{
                window.opener.location.href = "{parent_redirect}";
            }}
            setTimeout(() => {{ window.close(); }}, 100);
        </script></head>
        <body style="font-family:Arial,sans-serif;text-align:center;padding:40px;background:#f8fdf8">
            <div style="background:#d4edda;border:1px solid #c3e6cb;border-radius:8px;padding:25px;max-width:400px;margin:0 auto">
                <h2 style="color:#155724;margin-top:0">✅ Successfully Inserted</h2>
                <p style="color:#155724">Passport moved to valid records.</p>
            </div>
        </body></html>
        ''', 200

    except Exception as e:
        invalid_passport = get_invalid_passport_by_id(invalid_id, user_id)
        mrz_lines = [l.strip() for l in mrz_text.split("\n") if l.strip()]
        parsed, errors = parse_mrz(mrz_lines)
        if parsed is None:
            parsed = {}
        _error_msg = str(e) if isinstance(e, GroupVisaTypeMismatch) else f"Error during insert: {str(e)}"
        return render_template(
            "reparse.html",
            mrz_lines=mrz_lines, parsed=parsed,
            errors=[_error_msg],
            nationality_options=NATIONALITY_OPTIONS,
            marital_status_options=MARITAL_STATUS_OPTIONS,
            defaults=get_user_settings(user_id),
            NATIONALITY_CODE_MAP=NATIONALITY_CODE_MAP
        )


def generate_badge_image(passport_data, username="default", company_name=None):
    """
    company_name is what gets printed in the "company" slot on the badge.
    - username 'luxury' always shows "GROUP ABDU" regardless of what's passed in.
    - everyone else shows their account's full name (falls back to username
      if no company_name was supplied, e.g. old call sites).
    """
    if username == "luxury":
        company_name = "GROUP ABDU"
    elif not company_name:
        company_name = username

    user_template_path = f"{username}.jpg"
    template_path = user_template_path if os.path.exists(user_template_path) else "badge_template.jpg"
    try:
        template = Image.open(template_path)
        if template.mode != 'RGB':
            template = template.convert('RGB')
        draw = ImageDraw.Draw(template)
        try:
            font_bold = ImageFont.truetype(_resource_path("arialbd.ttf"), 40)
            font_regular = ImageFont.truetype(_resource_path("arial.ttf"), 40)
        except IOError:
            font_bold = font_regular = ImageFont.load_default()

        surname = passport_data.get('given_names', '') or ''
        given_names = passport_data.get('surname', '') or ''
        full_name = f"{surname} {given_names}".strip()
        passport_num = passport_data.get('passport_number', '') or ''
        nationality_id = passport_data.get('nationality_id')
        nationality_name = passport_data.get('country', '')
        for nid, nname in NATIONALITY_OPTIONS:
            if nid == nationality_id:
                nationality_name = nname
                break

        coords = {'name': (250, 700), 'passport': (400, 770), 'nationality': (370, 910), 'company': (330, 980)}
        draw.text(coords['name'], full_name, fill="black", font=font_regular)
        draw.text(coords['passport'], passport_num, fill="black", font=font_regular)
        draw.text(coords['nationality'], nationality_name, fill="black", font=font_regular)
        draw.text(coords['company'], company_name.upper(), fill="black", font=font_regular)

        # Load face from filesystem
        _face_fn = passport_data.get('filename')
        if _face_fn:
            _, _face_path = resolve_passport_paths(_face_fn, passport_data.get('group_name'))
            if _face_path and os.path.exists(_face_path):
                try:
                    face_img = Image.open(_face_path).resize((297, 300))
                    template.paste(face_img, (289, 370))
                except Exception as _fe:
                    print(f"Badge face load error: {_fe}")

        img_byte_arr = io.BytesIO()
        template.save(img_byte_arr, format='JPEG', quality=95)
        return img_byte_arr.getvalue()
    except Exception as e:
        print(f"Badge Generation Error: {e}")
        return None


def _generate_csv_response(passports_data, filename):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from collections import OrderedDict

    NAVY_MID = "112244"
    GOLD = "C9A84C"
    GOLD_LIGHT = "E8C97A"
    CREAM = "FAF8F4"
    TEXT_DARK = "0C1A30"
    BORDER_LIGHT = "E8E4DC"

    wb = Workbook()
    sheet = wb.active
    sheet.title = "Passport Data"
    headers = [
        'PASSPORT NO.', 'NATIONALITY', 'FIRST NAME', 'MIDDLE NAME', 'LAST NAME',
        'DATE OF BIRTH', 'SEX', 'ISSUE DATE', 'EXPIRY DATE', 'GROUP NAME'
    ]
    num_cols = len(headers)
    sheet.append(headers)

    thin_border = Border(*(Side(style='thin', color=BORDER_LIGHT),) * 4)
    header_fill = PatternFill('solid', fgColor=NAVY_MID)
    header_font = Font(bold=True, color="FFFFFF", size=11, name="DM Sans")
    stripe_fill = PatternFill('solid', fgColor=CREAM)
    gold_top_side = Side(style='medium', color=GOLD)
    group_fill = PatternFill('solid', fgColor=GOLD_LIGHT)
    group_font = Font(bold=True, color=NAVY_MID, size=13, name="DM Sans")

    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(top=gold_top_side, bottom=gold_top_side, left=thin_border.left, right=thin_border.right)
    sheet.row_dimensions[1].height = 22

    # Group rows by their group name, preserving first-seen order
    grouped = OrderedDict()
    for p in passports_data:
        gname = (p.get('group_name') or '').strip() or 'Ungrouped'
        grouped.setdefault(gname, []).append(p)

    nat_map = {nid: name for nid, name in NATIONALITY_OPTIONS}
    row_idx = 2
    for gname, group_rows in grouped.items():
        # Group title row, merged across all columns
        sheet.append([gname] + [''] * (num_cols - 1))
        sheet.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=num_cols)
        title_cell = sheet.cell(row=row_idx, column=1)
        title_cell.font = group_font
        title_cell.fill = group_fill
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        for col in range(1, num_cols + 1):
            sheet.cell(row=row_idx, column=col).fill = group_fill
            sheet.cell(row=row_idx, column=col).border = Border(
                top=gold_top_side, bottom=gold_top_side,
                left=thin_border.left, right=thin_border.right
            )
        sheet.row_dimensions[row_idx].height = 28
        row_idx += 1

        for stripe_pos, p in enumerate(group_rows):
            dob = p.get('dob')
            dob_str = dob.strftime('%d-%m-%Y') if dob else 'N/A'
            issue = p.get('passport_issue_date')
            issue_str = issue.strftime('%d-%m-%Y') if issue else 'N/A'
            expiry = p.get('expiry')
            expiry_str = expiry.strftime('%d-%m-%Y') if expiry else 'N/A'
            nat_name = nat_map.get(p.get('nationality_id'), '')
            country_code = p.get('country', '')
            nat_display = f"{nat_name} ({country_code})" if nat_name and country_code else str(nat_name or country_code)
            sheet.append([
                p.get('passport_number', ''), nat_display,
                p.get('given_names', ''), p.get('middle_name', ''), p.get('surname', ''),
                dob_str, p.get('sex', ''), issue_str, expiry_str, p.get('group_name', '')
            ])
            is_stripe = (stripe_pos % 2 == 1)
            for cell in sheet[row_idx]:
                cell.font = Font(color=TEXT_DARK, size=10.5, name="DM Sans")
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if is_stripe:
                    cell.fill = stripe_fill
            row_idx += 1

    sheet.freeze_panes = "A2"

    for col in range(1, num_cols + 1):
        col_letter = get_column_letter(col)
        max_len = 0
        for row in range(1, row_idx):
            val = sheet.cell(row=row, column=col).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        sheet.column_dimensions[col_letter].width = max_len + 4

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    if filename.lower().endswith('.csv'):
        filename = filename[:-4] + '.xlsx'
    output = make_response(bio.getvalue())
    output.headers["Content-Disposition"] = f"attachment; filename={filename}"
    output.headers["Content-type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return output


def _get_passports_by_group_names(user_id, group_names):
    """Return all passport rows for the given list of group names."""
    if not group_names:
        return []
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("USE passport_db")
    placeholders = ','.join(['%s'] * len(group_names))
    cursor.execute(f"""
        SELECT p.passport_number, p.country, p.surname, p.given_names, p.middle_name, p.dob, p.sex,
               g.passport_issue_date, p.expiry, g.group_name, g.nationality_id
        FROM passports p
        LEFT JOIN general_data g ON g.passport_id = p.id
        WHERE p.user_id = %s
          AND g.group_name IN ({placeholders})
          AND (p.is_recycled = FALSE OR p.is_recycled IS NULL)
        ORDER BY g.group_name, p.created_at
    """, [user_id] + list(group_names))
    data = cursor.fetchall()
    cursor.close()
    conn.close()
    return data


def _get_passport_ids_by_group_names(user_id, group_names):
    """Return passport IDs for the given group names."""
    if not group_names:
        return []
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("USE passport_db")
    placeholders = ','.join(['%s'] * len(group_names))
    cursor.execute(f"""
        SELECT p.id
        FROM passports p
        LEFT JOIN general_data g ON g.passport_id = p.id
        WHERE p.user_id = %s
          AND g.group_name IN ({placeholders})
          AND (p.is_recycled = FALSE OR p.is_recycled IS NULL)
    """, [user_id] + list(group_names))
    rows = cursor.fetchall()
    cursor.close()
    conn.close()
    return [r['id'] for r in rows]


@app.context_processor
def inject_globals():
    if 'user_id' in session:
        uid = session['user_id']
        try:
            recycled_total = (
                get_total_recycled_passports_count(uid) +
                get_total_recycled_invalid_count(uid)
            )
            total_invalid = get_total_invalid_count(uid)
        except Exception:
            _logger.exception(
                "inject_globals() context processor failed for user %s - "
                "falling back to zeros so the page still renders", uid
            )
            return {'total_invalid': 0, 'recycled_total': 0}
        return {
            'total_invalid': total_invalid,
            'recycled_total': recycled_total,
        }
    return {}
def _apply_issue_date_day_rule(db_date, expiry_str, country_code=None):
    """
    Sanity-checks an OCR-extracted issue date against the passport's expiry
    date, using only the DAY-OF-MONTH (month/year are ignored in the
    comparison itself).

    Rule:
      - If extracted_day == expiry_day  -> trust the OCR value as-is.
      - If they differ                 -> the OCR day is considered wrong.
        Replace ONLY the day, keeping the extracted date's own month/year,
        with:
            expiry_day + 1   (general passports)
            expiry_day + 0   (KGZ / RUS passports -- no +1 offset)

    Returns the (possibly corrected) date as a 'YYYY-MM-DD' string, or the
    original db_date unchanged if either date can't be parsed.

    Shared by app_routes.py (manual crop-scan issue-date routes) and
    reparse_routes.py (rescan flows) — lives here so both can import it
    via `from app_core import *`.
    """
    if not db_date or not expiry_str:
        return db_date

    try:
        extracted_dt = datetime.strptime(db_date, "%Y-%m-%d").date()
        expiry_dt = datetime.strptime(expiry_str, "%Y-%m-%d").date()
    except Exception:
        return db_date

    if extracted_dt.day == expiry_dt.day:
        return db_date  # OCR day matches expiry day -- trust it as-is

    # Mismatch: rebuild the day from expiry's day, +1 unless KGZ/RUS,
    # while keeping the extracted date's own month/year.
    offset = 0 if (country_code in ("KGZ", "RUS")) else 1
    corrected_day = expiry_dt.day + offset

    try:
        corrected_dt = extracted_dt.replace(day=corrected_day)
    except ValueError:
        # corrected_day overflows the extracted month's day count
        # (e.g. Feb 29/30/31) -- roll over into the next month instead
        # of silently failing.
        corrected_dt = extracted_dt.replace(day=1) + timedelta(days=corrected_day - 1)

    return corrected_dt.strftime("%Y-%m-%d")


@app.after_request
def apply_global_text_changes(response):
    """
    Globally replaces specific text in all HTML responses.
    Useful for updating navigation labels across all templates without editing each file.
    """
    # Only modify HTML responses (ignore JSON/API/XML)
    if response.content_type and response.content_type.startswith('text/html'):
        # response.data is bytes, so we replace bytes
        response.data = response.data.replace(b'View Results', b'View Records')
        
        # You can add more replacements here if needed:
        # response.data = response.data.replace(b'Old Text', b'New Text')
        
    return response